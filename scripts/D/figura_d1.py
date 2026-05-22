"""
Figura D.1 — Disponibilidad de las TIC en los hogares (2010-2023)
Fuente:
  - 2001-2014: INEGI MODUTIH  → 27_2023_hnal110.xlsx  (Computadora, Radio, Telefonía)
                               → 30_2023_hnal130.xlsx  (TV digital / analógico)
  - 2015-2023: INEGI ENDUTIH  → mismos archivos
"""

import openpyxl
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from pathlib import Path
import sys
import os

# Configuración de rutas para importar módulos locales
sys.path.append(str(Path(__file__).resolve().parents[1]))
try:
    from _plot_data_logger import enable_plot_data_logging
    enable_plot_data_logging()
except ImportError:
    pass

# Configuración global de tipografía requerida por la CRT
plt.rcParams['font.family'] = ['Noto Sans', 'DejaVu Sans', 'sans-serif']

# ── Rutas de entrada ────────────────────────────────────────────────────────
BASE = r"C:\Users\ivan-\Documents\GitHub\anuario\datos\D.1"
FILE27 = os.path.join(BASE, "27_2023_hnal110.xlsx")
FILE30 = os.path.join(BASE, "30_2023_hnal130.xlsx")

# ── Funciones de Lectura de Datos ──────────────────────────────────────────
def leer_archivo27(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        raw = str(row[0]).strip()
        anio_str = ''.join(c for c in raw if c.isdigit())
        if len(anio_str) == 4:
            anio = int(anio_str)
            data[anio] = {
                'computadora': row[2],
                'radio':       row[12],
                'telefonia':   row[10],
            }
    return data

def leer_archivo30(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        raw = str(row[0]).strip()
        anio_str = ''.join(c for c in raw if c.isdigit())
        if len(anio_str) == 4:
            anio = int(anio_str)
            solo_dig = row[4] or 0
            solo_ana = row[6] or 0
            ambos    = row[8] or 0
            data[anio] = {
                'tv_digital':   solo_dig + ambos,
                'tv_analogico': solo_ana + ambos,
            }
    return data

data27 = leer_archivo27(FILE27)
data30 = leer_archivo30(FILE30)

# ── Construir series 2010-2023 ──────────────────────────────────────────────
years = list(range(2010, 2024))

comp  = [round(data27[y]['computadora'] or 0) for y in years]
radio = [round(data27[y]['radio']       or 0) for y in years]
cel   = [round(data27[y]['telefonia']   or 0) for y in years]
tvdig = [round(data30[y]['tv_digital']  or 0) for y in years]
tvana = [round(data30[y]['tv_analogico']or 0) for y in years]

# ── Colores Institucionales CRT ─────────────────────────────────────────────
COLOR_CEL   = '#006157'   # Verde institucional EXACTO de la figura A.3
COLOR_TVDIG = '#b35aba'   # Morado (Paleta figura A.3)
COLOR_COMP  = '#ed8945'   # Naranja (Paleta operadores/guía)
COLOR_RADIO = '#368491'   # Azul medio (Paleta operadores/guía)
COLOR_TVANA = '#8e244d'   # Rosa-vino (Paleta operadores/guía)

# ── 1. Configuración de Figura ──────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(16, 8.5))
fig.patch.set_facecolor('white')
ax.set_facecolor('#F8F8FA')

x = np.arange(len(years))

# ── 2. Trazado de Líneas ────────────────────────────────────────────────────
lw = 2.5
ms = 6

def graficar_linea(eje_x, eje_y, color, label):
    return ax.plot(eje_x, eje_y, color=color, linewidth=lw, zorder=4,
                   marker='o', markersize=ms, markerfacecolor=color,
                   markeredgecolor=color, markeredgewidth=0,
                   label=label)[0]

l1 = graficar_linea(years, cel, COLOR_CEL, 'Teléfono celular')
l2 = graficar_linea(years, tvdig, COLOR_TVDIG, 'Televisor digital')
l3 = graficar_linea(years, comp, COLOR_COMP, 'Equipo de cómputo')
l4 = graficar_linea(years, radio, COLOR_RADIO, 'Aparatos de radio')
l5 = graficar_linea(years, tvana, COLOR_TVANA, 'Televisor analógico')

# ── 3. Algoritmo de Etiquetado por Clustering Espacial ──────────────────────
def etiquetar_dinamico_por_clusters(ax, years, all_series_data):
    def create_bbox(color):
        return dict(boxstyle='round,pad=0.3,rounding_size=0.8',
                    facecolor='white', edgecolor=color, linewidth=0.8)
    
    for i, year in enumerate(years):
        # 1. Recopilar y ordenar los 5 valores de este año de mayor a menor
        pts = [{'val': data['values'][i], 'color': data['color']} for data in all_series_data]
        pts.sort(key=lambda x: x['val'], reverse=True)
        
        # 2. Agrupar puntos que estén demasiado cerca (<= 9 unidades de diferencia)
        clusters = []
        current_cluster = [pts[0]]
        
        for j in range(1, len(pts)):
            if current_cluster[-1]['val'] - pts[j]['val'] <= 9:
                current_cluster.append(pts[j])
            else:
                clusters.append(current_cluster)
                current_cluster = [pts[j]]
        clusters.append(current_cluster)
        
        # 3. Asignar posiciones geométricas basadas en el tamaño del cluster
        for cluster in clusters:
            n = len(cluster)
            
            if n == 1:
                # Punto aislado: Lo mandamos arriba o abajo dependiendo de su valor para no tapar la línea
                cluster[0]['x_off'] = 0
                cluster[0]['y_off'] = 14 if cluster[0]['val'] >= 50 else -14
                cluster[0]['va'] = 'bottom' if cluster[0]['val'] >= 50 else 'top'
                
            elif n == 2:
                # Dos puntos muy juntos: Uno arriba y uno abajo
                cluster[0]['x_off'], cluster[0]['y_off'], cluster[0]['va'] = 0, 14, 'bottom'
                cluster[1]['x_off'], cluster[1]['y_off'], cluster[1]['va'] = 0, -14, 'top'
                
            elif n == 3:
                # Tres puntos juntos (Ej. año 2013): Arriba, izquierda-centro, abajo
                cluster[0]['x_off'], cluster[0]['y_off'], cluster[0]['va'] = 0, 15, 'bottom'
                cluster[1]['x_off'], cluster[1]['y_off'], cluster[1]['va'] = -18, 0, 'center'
                cluster[2]['x_off'], cluster[2]['y_off'], cluster[2]['va'] = 0, -15, 'top'
                
            elif n == 4:
                # Cuatro puntos: Arriba, izquierda-arriba, derecha-abajo, abajo
                cluster[0]['x_off'], cluster[0]['y_off'], cluster[0]['va'] = 0, 16, 'bottom'
                cluster[1]['x_off'], cluster[1]['y_off'], cluster[1]['va'] = -18, 6, 'bottom'
                cluster[2]['x_off'], cluster[2]['y_off'], cluster[2]['va'] = 18, -6, 'top'
                cluster[3]['x_off'], cluster[3]['y_off'], cluster[3]['va'] = 0, -16, 'top'
                
            elif n >= 5:
                # Todos amontonados: Patrón de estrella completo
                cluster[0]['x_off'], cluster[0]['y_off'], cluster[0]['va'] = 0, 18, 'bottom'
                cluster[1]['x_off'], cluster[1]['y_off'], cluster[1]['va'] = -20, 8, 'bottom'
                cluster[2]['x_off'], cluster[2]['y_off'], cluster[2]['va'] = 20, 0, 'center'
                cluster[3]['x_off'], cluster[3]['y_off'], cluster[3]['va'] = -20, -8, 'top'
                cluster[4]['x_off'], cluster[4]['y_off'], cluster[4]['va'] = 0, -18, 'top'
                
        # 4. Dibujar las etiquetas en el gráfico
        for p in pts:
            bbox = create_bbox(p['color'])
            ax.annotate(f"{p['val']}", xy=(year, p['val']),
                        xytext=(p['x_off'], p['y_off']), textcoords='offset points',
                        ha='center', va=p['va'], fontsize=8, fontweight='bold',
                        color='#3c3c3b', bbox=bbox, zorder=10)

all_data = [
    {'values': cel,   'color': COLOR_CEL},
    {'values': tvdig, 'color': COLOR_TVDIG},
    {'values': comp,  'color': COLOR_COMP},
    {'values': radio, 'color': COLOR_RADIO},
    {'values': tvana, 'color': COLOR_TVANA},
]

etiquetar_dinamico_por_clusters(ax, years, all_data)

# ── 4. Formato de Ejes ──────────────────────────────────────────────────────
ax.set_ylim(0, 110)
ax.yaxis.set_major_locator(mticker.MultipleLocator(10))
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{int(v)}%'))
ax.tick_params(axis='y', labelsize=10, colors='#3c3c3b') 

for label in ax.get_yticklabels():
    label.set_fontweight('medium')

ax.set_xticks(years)
ax.set_xticklabels([str(y) for y in years], fontsize=10, color='#3c3c3b', fontweight='bold')
ax.tick_params(axis='x', length=0, color='#7c7c7c', pad=8)

ax.grid(axis='y', color='#d1d1d1', linewidth=1, zorder=0)

for spine in ax.spines.values():
    spine.set_color('#7c7c7c')
    spine.set_linewidth(1)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

# ── 5. Encabezado de la Figura ──────────────────────────────────────────────
ax.annotate(' ', xy=(0, 1), xycoords='axes fraction',
            xytext=(0, 34), textcoords='offset points',
            fontsize=2, bbox=dict(boxstyle='round,pad=1.6,rounding_size=0.2', facecolor='#4a7d75', edgecolor='none'))

ax.annotate('Figura D.1.', xy=(0, 1), xycoords='axes fraction',
            xytext=(15, 30), textcoords='offset points',
            fontsize=14, fontweight='bold', color='#3c3c3b')

ax.annotate('Disponibilidad de las TIC en los hogares (2010-2023)', xy=(0, 1), xycoords='axes fraction',
            xytext=(100, 30), textcoords='offset points',
            fontsize=14, fontweight='medium', color='#3c3c3b')

# ── 6. Leyenda ──────────────────────────────────────────────────────────────
fig.legend(handles=[l1, l2, l3, l4, l5], loc='lower center', 
           bbox_to_anchor=(0.5, 0.12), ncol=5,
           frameon=False, handlelength=2.5, labelcolor='#3c3c3b',
           prop={'weight': 'bold', 'size': 10})

# ── 7. Notas al pie ─────────────────────────────────────────────────────────
ax.annotate('Fuente:', xy=(0.08, 0.06), xycoords='figure fraction',
            fontsize=8, fontweight='bold', color='#3c3c3b')
ax.annotate('IFT con datos del MODUTIH para el periodo 2010-2014 y la ENDUTIH para el periodo 2015-2023, del INEGI.',
            xy=(0.08, 0.06), xycoords='figure fraction',
            xytext=(35, 0), textcoords='offset points',
            fontsize=8, fontweight='normal', color='#3c3c3b')

# ── 8. Guardar ──────────────────────────────────────────────────────────────
fig.subplots_adjust(left=0.08, right=0.92, top=0.85, bottom=0.22)

out = 'output/Figura_D1.png'
os.makedirs(os.path.dirname(out), exist_ok=True)
fig.savefig(out, dpi=200, bbox_inches='tight', facecolor='white', edgecolor='none')
print(f"\nGráfica guardada en: {out}")
plt.close(fig)