"""
Figura A.5 — Inversión Extranjera Directa (IED) en telecomunicaciones

Gráfica de barras horizontales:
  - Barra larga (azul claro -> ahora teal claro): IED total de México (millones de dólares)
  - Barra corta (púrpura oscuro -> ahora teal oscuro): IED en Telecomunicaciones (sector 517 SCIAN)

Datos actualizados al 3er trimestre de 2025.
Período: 2013-2024 (2024 acumulado a junio).

Fuente: Secretaría de Economía – Registro Nacional de Inversiones Extranjeras.
"""

import os
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from pathlib import Path
import sys

try:
    sys.path.append(str(Path(__file__).resolve().parents[1]))
    from _plot_data_logger import enable_plot_data_logging
    enable_plot_data_logging()
except ImportError:
    pass

import matplotlib.ticker as mticker

# ── 1. Leer datos ─────────────────────────────────────────────────────────────
base = os.path.join(os.path.dirname(__file__), "..", "..", 'datos', 'A.5')

# --- IED total de México (datos actualizados) ---
wb1 = openpyxl.load_workbook(
    os.path.join(base, 'Datos_originales_y_actualizacion__1_.xlsx'),
    data_only=True)
ws1 = wb1['Preliminares y actualización']

total_ied = {}
for r in range(3, ws1.max_row + 1):
    yr = ws1.cell(r, 1).value
    period = ws1.cell(r, 2).value
    val = ws1.cell(r, 4).value  # Columna D = datos actualizados
    if yr and period and val:
        yr = int(yr)
        if 2013 <= yr <= 2024:
            if yr < 2024 and 'diciembre' in str(period):
                total_ied[yr] = float(val)
            elif yr == 2024 and 'junio' in str(period):
                total_ied[yr] = float(val)
wb1.close()

# --- IED en telecomunicaciones (sector 517, datos actualizados) ---
wb2 = openpyxl.load_workbook(
    os.path.join(base, '2025_3T_Flujosportipodeinversion_actu__3_.xlsx'),
    data_only=True, read_only=True)
ws2 = wb2['Por sector']

YEAR_START = 2006
telecom_ied = {}
for row in ws2.iter_rows(min_row=5, max_col=80, values_only=False):
    cell_a = str(row[0].value) if row[0].value else ''
    if cell_a.startswith('517 '):
        for yr in range(2013, 2025):
            if yr < 2024:
                idx = (yr - YEAR_START) * 4 + 3   # Q4 = anual
            else:
                idx = (yr - YEAR_START) * 4 + 1   # Q2 = enero-junio 2024
            v = row[idx + 1].value  # +1 porque row[0] es el label
            if v is not None and str(v) != 'C':
                telecom_ied[yr] = float(v)
            else:
                telecom_ied[yr] = 0.0
        break
wb2.close()

# ── 2. Preparar arrays ────────────────────────────────────────────────────────
years = list(range(2013, 2025))
ied_mexico = np.array([total_ied[y] for y in years])
ied_telecom = np.array([telecom_ied[y] for y in years])

# Imprimir tabla de verificación
print(f"{'Año':<6} {'IED México':>14} {'IED Telecom':>14}")
print("-" * 36)
for i, yr in enumerate(years):
    print(f"{yr:<6} {ied_mexico[i]:>14,.2f} {ied_telecom[i]:>14,.2f}")

# ── 3. Gráfica ESTILO A.9 ─────────────────────────────────────────────────────
plt.rcParams['font.family'] = ['Noto Sans', 'DejaVu Sans', 'sans-serif']

fig, ax = plt.subplots(figsize=(16, 8.5))
fig.patch.set_facecolor('white')
ax.set_facecolor('#F8F8FA')

n = len(years)
y_pos = np.arange(n)
bar_height = 0.38

# Colores replicados de Figura A.9
COLOR_MEXICO = '#86adae'     # Teal claro (sustituye a Azul Claro)
COLOR_TELECOM = '#335a5c'    # Teal oscuro (sustituye a Púrpura Oscuro)
color_texto = '#3c3c3b'

# Barras horizontales
bars_mx = ax.barh(y_pos - bar_height / 2, ied_mexico, height=bar_height,
                  color=COLOR_MEXICO, edgecolor='none', zorder=2,
                  label='Inversión Extranjera Directa de México')

bars_tc = ax.barh(y_pos + bar_height / 2, ied_telecom, height=bar_height,
                  color=COLOR_TELECOM, edgecolor='none', zorder=2,
                  label='Inversión Extranjera Directa en Telecomunicaciones')

# ── 4. Anotaciones de valor ───────────────────────────────────────────────────
for i, yr in enumerate(years):
    mx_val = ied_mexico[i]
    ax.text(mx_val + 300, y_pos[i] - bar_height / 2, f'{mx_val:,.0f}'.replace(',', ','),
            va='center', ha='left', fontsize=9, color=color_texto, fontweight='normal', zorder=3)

    tc_val = ied_telecom[i]
    if tc_val >= 0:
        ax.text(tc_val + 300, y_pos[i] + bar_height / 2, f'{tc_val:,.2f}'.replace(',', ','),
                va='center', ha='left', fontsize=9, color=color_texto, fontweight='normal', zorder=3)
    else:
        ax.text(tc_val - 300, y_pos[i] + bar_height / 2, f'{tc_val:,.2f}'.replace(',', ','),
                va='center', ha='right', fontsize=9, color=color_texto, fontweight='normal', zorder=3)

# ── 5. Ejes ───────────────────────────────────────────────────────────────────
ax.set_yticks(y_pos)
ax.set_yticklabels(years, fontsize=9, fontweight='normal', color=color_texto)
ax.invert_yaxis()  # Mantenemos lógica de orden del año original de A.5

ax.set_xlabel('Millones de dólares', fontsize=11, fontweight='medium', color=color_texto, labelpad=15)
ax.set_ylabel('Año', fontsize=11, fontweight='medium', color=color_texto, labelpad=15)

# Rango del eje X
x_min = min(ied_telecom.min(), 0) - 4000
x_max = ied_mexico.max() + 5000
ax.set_xlim(x_min, x_max)
ax.xaxis.set_major_locator(mticker.MultipleLocator(10000))
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{int(v):,}'.replace(',', ',')))
ax.tick_params(axis='x', labelsize=9, colors=color_texto)

# Grid y bordes A.9
ax.grid(axis='x', alpha=1.0, color='#d1d1d1', linewidth=1, zorder=0)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.spines['bottom'].set_color('#7c7c7c')
ax.spines['left'].set_color('#7c7c7c')

# Lína vertical en 0 (necesaria en esta gráfica por valores negativos)
ax.axvline(x=0, color='#999999', linewidth=0.8, zorder=1) 

# ── 6. Títulos (Bloque Institucional A.9) ─────────────────────────────────────
ax.annotate('   ', xy=(0, 1), xycoords='axes fraction', 
            xytext=(0, 30), textcoords='offset points',
            va='center', ha='left', fontsize=2,
            bbox=dict(boxstyle='round,pad=1.6,rounding_size=0.2', facecolor='#4a7d75', edgecolor='none'))

ax.annotate("Figura A.5.", xy=(0, 1), xycoords='axes fraction', 
            xytext=(15, 30), textcoords='offset points',
            fontsize=14, fontweight='bold', color=color_texto, ha='left', va='center')

ax.annotate(" Inversión Extranjera Directa (IED) en telecomunicaciones", 
            xy=(0, 1), xycoords='axes fraction', 
            xytext=(95, 30), textcoords='offset points',
            fontsize=14, fontweight='medium', color=color_texto, ha='left', va='center')

# ── 7. Leyenda ────────────────────────────────────────────────────────────────
handles, labels = ax.get_legend_handles_labels()
fig.legend(handles, labels, loc='lower center', bbox_to_anchor=(0.5, 0.08), ncol=2, fontsize=10, frameon=False, handlelength=2.5)

# ── 8. Notas al pie ───────────────────────────────────────────────────────────
font_size_notes = 8
x_start = 0.08

# Fuente
y_fuente = 0.06
ax.annotate("Fuente: ", xy=(x_start, y_fuente), xycoords='figure fraction',
            fontsize=font_size_notes, fontweight='bold', color=color_texto, ha='left', va='top')
note_fuente = 'IFT con datos de la Secretaría de Economía (datos actualizados al 3er trimestre de 2025). Datos disponibles en: https://www.gob.mx/se/acciones-y-programas/...'
ax.annotate(note_fuente, xy=(x_start, y_fuente), xycoords='figure fraction',
            xytext=(35, 0), textcoords='offset points',
            fontsize=font_size_notes, fontweight='normal', color=color_texto, ha='left', va='top')

# Notas
y_nota = 0.042
ax.annotate("Notas: ", xy=(x_start, y_nota), xycoords='figure fraction',
            fontsize=font_size_notes, fontweight='bold', color=color_texto, ha='left', va='top')
note_notas = 'Cifras en millones de dólares (dólares corrientes). Rama 5151 Transmisión de programas de radio y TV, Subsector 517 Telecomunicaciones. 2024 acumulada a junio.'
ax.annotate(note_notas, xy=(x_start, y_nota), xycoords='figure fraction',
            xytext=(32, 0), textcoords='offset points',
            fontsize=font_size_notes, fontweight='normal', color=color_texto, ha='left', va='top')

# ── 9. Guardar ────────────────────────────────────────────────────────────────
fig.subplots_adjust(left=0.08, right=0.92, top=0.85, bottom=0.22)
output_path = os.path.join(os.path.dirname(__file__), "..", "..", 'output', 'Figura_A5.png')
fig.savefig(output_path, dpi=200, bbox_inches='tight', facecolor='white', edgecolor='none')
print(f"\nGráfica guardada en: {output_path}")
plt.close(fig)